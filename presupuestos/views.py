from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.views.generic.base import TemplateView  
from .filters import ArticulosFilter
from .form import ConsForm, ArticulosForm
from proyectos.models import Proyectos
from computos.models import Computos
from compras.models import Compras
from ventas.models import PricingResumen, VentasRealizadas
from registro.models import RegistroValorProyecto
from .models import Articulos, Constantes, DatosProyectos, Prametros, Desde, Analisis, CompoAnalisis, Modelopresupuesto, Capitulos
import sqlite3
import numpy as np
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# Vistas para articulos desde linea 26 a 200
# Vistas para constantes desde linea 200 a 300






# --------------------------------> VISTA PARA LISTADO DE ARTICULOS <------------------------------------------------------


def insum_list(request):

    datos = Articulos.objects.all()

    #Aqui empieza el filtro

    if request.method == 'POST':

        palabra_buscar = request.POST.items()

        datos_viejos = datos

        datos = []   

        for i in palabra_buscar:

            if i[0] == "palabra":
        
                palabra_buscar = i[1]

        if str(palabra_buscar) == "":

            datos = datos_viejos

        else:
        
            for i in datos_viejos:

                palabra =(str(palabra_buscar))

                codigo = (str(i.codigo))

                nombre = (str(i.nombre))

                constante = (str(i.constante))

                valor = (str(i.valor))


                if palabra.lower() in codigo.lower() or palabra.lower() in nombre.lower() or palabra.lower() in constante.lower() or palabra.lower() in valor.lower():

                    datos.append(i)


    #Aqui termina el filtro

    c = {'datos':datos}

    return render(request, 'articulos/insum_list.html', c )

# --------------------------------> VISTA PARA PANEL DE MODIFICIACIÓN DE ARTICULOS <------------------------------------------------------


def insum_panel(request):

    art_actuales = Articulos.objects.all()

    myfilter = ArticulosFilter(request.GET, queryset=art_actuales)

    art_actuales = myfilter.qs

    c = {'articulos':art_actuales, 'myfilter':myfilter}

    return render(request, 'articulos/insum_panel.html', c )

# ----------------------------------> VISTAS PARA CREAR ARTICULOS <----------------------------------------------
    
def insum_create(request):

    #Si el metodo es POST activa la funciones para guardar los datos del formulario

    mensaje = ""

    if request.method == 'POST':

        try:

            #Aqui guardo los datos para ingresar el formulario

            form = ArticulosForm(request.POST)

            #Aqui guardo los datos de los inputs

            datos = request.POST.items()

            for key, value in datos:

                if key == 'codigo':

                    #Aqui solamente me quedo con el codigo
                    codigo = (value)

                if key == 'constante':

                    #Aqui solamente me quedo con el codigo
                    constante = (value)

                if key == 'valor':

                    #Aqui solamente me quedo con el codigo
                    valor = (value)

            #Aqui pruebo si el formulario es correcto

            if form.is_valid():
                
                form.save()
            
            #Me conecto a la base de datos y traigo el valor de la constante

            objetos_constante = Constantes.objects.all()

            for i in objetos_constante:

                if float(i.id) == float(constante):

                    valor_constante = float(i.valor)

                    #Opero para sacar el valor auxiliar 

                    valor_aux = (float(valor)/valor_constante)

                    objetos_insumos = Articulos.objects.all()

                    for i in objetos_insumos:

                        if int(i.codigo) == int(codigo):

                            i.valor_aux = valor_aux

                            i.save()

                            return redirect('Panel de cambios')
        except:

             mensaje = "Hay un error al cargar, cuidado con los puntos y comas"   
    else:
        form = ArticulosForm()

    f = {'form':form, 'mensaje':mensaje}

    return render(request, 'articulos/insum_create.html', f )

# --------------------------------> VISTA PARA EDITAR ARTICULOS <------------------------------------------------------

def insum_edit(request, id_articulos):

    art = Articulos.objects.get(codigo=id_articulos)

    if request.method == 'GET':
        form = ArticulosForm(instance = art)
    else:
        form = ArticulosForm(request.POST, instance = art)
        if form.is_valid():
            form.save()
        return redirect('Panel de cambios')

    return render(request, 'articulos/insum_create.html', {'form':form})

# --------------------------------> VISTA PARA CONFIRMAR SI SE ELIMINA UN ARTICULO <------------------------------------------------------

def insum_delete(request, id_articulos):

    art = Articulos.objects.get(codigo=id_articulos)

    if request.method == 'POST':
        art.delete()
        return redirect('Panel de cambios')

    return render(request, 'articulos/insum_delete.html', {'art':art})

# ----------------------------------------------------- VISTAS PARA CONSTANTES ----------------------------------------------

# VISTA --> Crear constante

def cons_create(request):

    if request.method == 'POST':
        form = ConsForm(request.POST)
        if form.is_valid():
            form.save()
        return redirect('Cons_panel')
        
    else:
        form = ConsForm()

    f = {'form':form}
    return render(request, 'constantes/cons_create.html', f )

# ----------------------------------------------------- VISTAS PARA PANEL DE CAMBIOS CONSTANTES ----------------------------------------------

def cons_list(request):

    cons_actuales = Constantes.objects.all()

    c = {'constantes':cons_actuales}

    return render(request, 'constantes/cons_list.html', c )

def cons_panel(request):

    cons_actuales = Constantes.objects.all()

    c = {'constantes':cons_actuales}

    return render(request, 'constantes/cons_panel.html', c )

# --------------------------------> VISTA PARA EDITAR CONSTANTES <------------------------------------------------------

def cons_edit(request, id_cons):

    cons = Constantes.objects.get(id=id_cons)

    if request.method == 'GET':
        form = ConsForm(instance = cons)
    else:
        form = ConsForm(request.POST, instance = cons)
        
        if form.is_valid():

            # Rescato el nombre y el valor nuevo de la constante

            datos = request.POST.items() 

            for key, value in datos:

                if key == 'nombre':

                    nombre = (value)
                    
                if key == 'valor':

                    cons_valor_nuevo = (value)

            datos_constante = Constantes.objects.all()

            for i in datos_constante:
                if str(i.nombre) == str(nombre):
                    
                    cons_nombre = i.nombre

                    cons_valor = i.valor

                    form.save()

            datos_insumos = Articulos.objects.all()

            for i in datos_insumos:

                if str(i.constante) == str(cons_nombre):
                    valor_actual = i.valor

                    valor_nuevo = valor_actual*(float(cons_valor_nuevo)/cons_valor) 

                    i.valor = valor_nuevo

                    i.save()

        return redirect('Cons_panel')
    
    return render(request, 'constantes/cons_create.html', {'form':form})

# --------------------------------> VISTA PARA ELIMINAR CONSTANTE <------------------------------------------------------

def cons_delete(request, id_cons):

    cons = Constantes.objects.get(id=id_cons)

    if request.method == 'POST':
        cons.delete()
        return redirect('Cons_panel')
    return render(request, 'constantes/cons_delete.html', {'cons':cons})

# ---------------------------------> VISTAS PARA PANEL PRESUPUESTOS <----------------------------------------------

def presupuestostotal(request):
    
    proyectos = Proyectos.objects.all()

    datos = 0

    registro = 0

    if request.method == 'POST':

        #Trae el proyecto elegido

        proyecto_elegido = request.POST.items()

        #Crea los datos del pricing

        for i in proyecto_elegido:

            if i[0] == "proyecto":
                proyectos = Proyectos.objects.get(id = i[1])

        datos = []

        datos_presupuesto = PresupuestoPorCapitulo(proyectos.id)
        datos_saldo = Saldoporcapitulo(proyectos.id)

        valor_reposicion = 0

        for p in datos_presupuesto:

            for articulo_cantidad in p[2]:

                valor_reposicion = (valor_reposicion + articulo_cantidad[0].valor*articulo_cantidad[1])

        valor_reposicion = valor_reposicion/1000000
        
        valor_saldo = 0

        for s in datos_saldo:

            for articulo_cantidad in s[2]:

                if articulo_cantidad[1] > 0:

                    valor_saldo = (valor_saldo + articulo_cantidad[0].valor*articulo_cantidad[1])

        valor_saldo = valor_saldo/1000000
        avance = 0

        if valor_reposicion != 0:
            avance = (1 - (valor_saldo/valor_reposicion))*100

        datos.append((proyectos, valor_reposicion, valor_saldo, avance))

        valor_proyecto = RegistroValorProyecto.objects.filter(proyecto = proyectos)

        registro = []

        for valor in valor_proyecto:


            registro.append((valor.fecha, valor.precio_proyecto/1000000))

        proyectos = 0
        

    return render(request, 'presupuestos/principalpresupuesto.html', {"datos":datos, "proyectos":proyectos, "valor":registro,})


# ---------------------------------> VISTAS PARA PANEL PRESUPUESTOS - SALDO CAPITULO ----------------------------------------------

def saldocapitulo(request, id_proyecto):

    #Armamos los datos para ver el presupuesto por capitulo

    datos = PresupuestoPorCapitulo(id_proyecto)

    datos_viejos = datos

    datos_presupuesto = []

    for componentes in datos_viejos:

        valor_capitulo = 0

        for articulos in componentes[2]:

            valor_capitulo = valor_capitulo + articulos[0].valor*articulos[1]
        
        datos_presupuesto.append((componentes[0], componentes[1], valor_capitulo ))


    #Armamos el saldo de cada capitulo

    saldo = Saldoporcapitulo(id_proyecto)

    datos_viejos = saldo

    datos_saldo = []

    valor_saldo = 0

    for componentes in datos_viejos:

        saldo_capitulo = 0

        for articulos in componentes[2]:

            if articulos[1] > 0:

                saldo_capitulo = saldo_capitulo + articulos[0].valor*articulos[1]
        
        datos_saldo.append((componentes[0], componentes[1], saldo_capitulo ))

        valor_saldo = valor_saldo + saldo_capitulo

    #Combinamos ambos

    datos = []

    for p in datos_presupuesto:
        for s in datos_saldo:
            if p[0] == s[0]:

                avance = 0
                inc = 0

                if p[2] != 0:

                    avance = (1 - s[2]/p[2])*100               
                    inc = (s[2]/valor_saldo)*100  

                datos.append((p[0], p[1], p[2], s[2], avance, inc))

    
    proyecto = Proyectos.objects.get(id = id_proyecto)

    datos = {"proyecto":proyecto, "datos":datos, "saldo":valor_saldo}
                
    return render(request, 'presupuestos/saldocapitulo.html', {"datos":datos})

# ----------------------------------------------------- VISTAS PARA ARTICULOS SALDO - CAPITULO ----------------------------------------------

def SaldoCapArticulos(request, id_proyecto, id_capitulo):

    #Armamos el saldo de cada capitulo

    saldo = Saldoporcapitulo(id_proyecto)

    datos_viejos = saldo

    datos_saldo = []
    capitulo = []

    for componentes in datos_viejos:
        if int(componentes[1].id) == int(id_capitulo):
            
            datos_saldo.append(componentes[2])
            capitulo.append(componentes[1])

    articulos = []

    for articulo in datos_saldo[0]:
        articulos.append(articulo[0])

    articulos = list(set(articulos))

    articulos_cant = []

    for articulo in articulos:

        cantidad = 0

        for art_can in datos_saldo[0]:

            if articulo == art_can[0] and art_can[1]>0:
                cantidad = cantidad + art_can[1]
        articulos_cant.append((articulo, cantidad))
    
    saldo_cap = 0

    datos_viejos = articulos_cant
    datos_saldo = []

    for dato in datos_viejos:
        saldo_cap = saldo_cap + dato[0].valor*dato[1]
        datos_saldo.append((dato[0], dato[1], float(dato[0].valor*dato[1])))

    datos_viejos = datos_saldo
    datos_saldo = []

    for dato in datos_viejos:
        if saldo_cap != 0:
            inc = float(dato[2])/float(saldo_cap)*100
        else:
            inc = 0
        datos_saldo.append((dato[0], dato[1], dato[2], inc))


    if len(datos_saldo) == 0:
        datos_saldo = 0

    else:

        datos_saldo.sort(key=lambda tup: tup[3], reverse=True)

    proyecto = Proyectos.objects.get(id = id_proyecto)

    datos = {"proyecto":proyecto,
     "datos_saldo":datos_saldo, "capitulo":capitulo,
     "saldo":saldo_cap}

    return render(request, 'presupuestos/saldoartcapitulo.html', {"datos":datos})

# ----------------------------------------------------- VISTAS PARA ARTICULOS SALDO - CAPITULO ----------------------------------------------

def debugsa(request, id_proyecto):

    datos = debugsaldo(id_proyecto)

    return render(request, 'presupuestos/debugsaldo.html', {"datos":datos})


# ----------------------------------------------------- VISTAS PARA PANEL PRESUPUESTOS - EXPLOSION ----------------------------------------------
def explosion(request, id_proyecto):

    proyecto = Proyectos.objects.get(id = id_proyecto)
    modelo = Modelopresupuesto.objects.filter(proyecto = proyecto)

    #Version 2 de explosión
    
    crudo_analisis = []

    for i in modelo:

        if i.cantidad != None:

            crudo_analisis.append((i.analisis, i.cantidad))

        else:

            if "SOLO MANO DE OBRA" in str(i.analisis.nombre):

                computos = Computos.objects.filter(tipologia = i.vinculacion, proyecto = proyecto)

                cantidad = 0

                for r in computos:
                    cantidad = cantidad + r.valor_vacio

                crudo_analisis.append((i.analisis, cantidad))

            else:

                computos = Computos.objects.filter(tipologia = i.vinculacion, proyecto = proyecto)

                cantidad = 0

                for r in computos:
                    cantidad = cantidad + r.valor_lleno

                crudo_analisis.append((i.analisis, cantidad))

    crudo_articulos = []


    for c in crudo_analisis:

        analisis = CompoAnalisis.objects.filter(analisis = c[0])

        for d in analisis:

            cantidad = d.cantidad*c[1]

            crudo_articulos.append((d.articulo, cantidad))

    datos = []

    for t in crudo_articulos:
        datos.append(t[0])

    datos = list(set(datos))

    datos_viejos = datos
    datos = []

    for i in datos_viejos:
        cantidad = 0
        for c in crudo_articulos:
            if i == c[0]:
                cantidad = cantidad + c[1]
        datos.append((i, cantidad))


    compras = Compras.objects.filter(proyecto = proyecto)

    datos_viejos = datos
    datos = []

    for i in datos_viejos:
        comprado = 0
        for c in compras:
            if c.articulo == i[0]:
                comprado = comprado + c.cantidad
        
        cantidad_saldo = i[1] - comprado

        saldo = cantidad_saldo * i[0].valor
        
        datos.append((i[0], i[1], comprado, cantidad_saldo, saldo ))

          #Aqui empieza el filtro

    if request.method == 'POST':

        palabra_buscar = request.POST.items()

        datos_viejos = datos

        datos = []   

        for i in palabra_buscar:

            if i[0] == "palabra":
        
                palabra_buscar = i[1]

        if str(palabra_buscar) == "":

            datos = datos_viejos

        else:
        
            for i in datos_viejos:

                palabra =(str(palabra_buscar))

                buscador = (str(i[0]))

                if palabra.lower() in buscador.lower():

                    datos.append(i)


    #Aqui termina el filtro

    datos = {"datos":datos,
    "proyecto":proyecto}

    return render(request, 'presupuestos/explosion.html', {"datos":datos})

# ----------------------------------------------------- VISTAS PARA PANEL PRESUPUESTOS - ANALISIS ----------------------------------------------
def presupuestosanalisis(request, id_proyecto, id_capitulo):

    proyecto = Proyectos.objects.get(id = id_proyecto)
    capitulo = Capitulos.objects.get(id = id_capitulo)
    compo = CompoAnalisis.objects.all()
    computo = Computos.objects.all()
    modelo = Modelopresupuesto.objects.filter(proyecto = proyecto)

    crudo = []

    valor_capitulo = 0

    for d in modelo:

        if d.capitulo == capitulo and d.proyecto == proyecto:

            if d.cantidad == None:

                if "SOLO MANO DE OBRA" in str(d.analisis):

                    valor_analisis = 0

                    for e in compo:

                        if e.analisis == d.analisis:

                            valor_analisis = valor_analisis + e.articulo.valor*e.cantidad

                    cantidad = 0

                    for h in computo:
                        if h.proyecto == proyecto and h.tipologia == d.vinculacion:
                            cantidad = cantidad + h.valor_vacio  

                    total_parcial = valor_analisis*cantidad/1000
                    crudo.append((d.analisis, valor_analisis, cantidad, total_parcial, 0.0)) 

                    valor_capitulo = valor_capitulo + valor_analisis*cantidad

                else:
                    valor_analisis = 0

                    for e in compo:

                        if e.analisis == d.analisis:

                            valor_analisis = valor_analisis + e.articulo.valor*e.cantidad

                    cantidad = 0

                    for h in computo:
                        if h.proyecto == proyecto and h.tipologia == d.vinculacion:
                            cantidad = cantidad + h.valor_lleno 

                    total_parcial = valor_analisis*cantidad/1000
                    crudo.append((d.analisis, valor_analisis, cantidad, total_parcial, 0.0, d.vinculacion)) 

                    valor_capitulo = valor_capitulo + valor_analisis*cantidad
            else:

                valor_analisis = 0

                for e in compo:

                    if e.analisis == d.analisis:

                        valor_analisis = valor_analisis + e.articulo.valor*e.cantidad
                
                total_parcial = valor_analisis*float(d.cantidad)/1000
                
                crudo.append((d.analisis, valor_analisis, d.cantidad, total_parcial, 0.0, d.vinculacion))
                
                valor_capitulo = valor_capitulo + valor_analisis*float(d.cantidad)
    datos =[]

    for i in crudo:
        i = list(i)
        i[4] = i[3]/valor_capitulo*100000
        i = tuple(i)
        datos.append(i)

    datos = {"datos":datos, "proyecto":proyecto, "capitulo":capitulo}

    return render(request, 'presupuestos/presupuestoanalisis.html', {"datos":datos})


# ----------------------------------------------------- VISTAS PARA PANEL PRESUPUESTOS - CAPITULO ----------------------------------------------

def presupuestoscapitulo(request, id_proyecto):

    proyecto = Proyectos.objects.get(id = id_proyecto)
    capitulo = Capitulos.objects.all()
    compo = CompoAnalisis.objects.all()
    computo = Computos.objects.all()
    modelo = Modelopresupuesto.objects.filter(proyecto = proyecto)

    if len(modelo) == 0:
        datos = 0

        datos = {"datos":datos, "proyecto":proyecto}
    
    else:

        crudo = []

        valor_proyecto = 0

        for c in capitulo:

            valor_capitulo = 0

            for d in modelo:

                if d.capitulo == c and d.proyecto == proyecto:

                    if d.cantidad == None:

                        if "SOLO MANO DE OBRA" in str(d.analisis):

                            valor_analisis = 0

                            for e in compo:

                                if e.analisis == d.analisis:

                                    valor_analisis = valor_analisis + e.articulo.valor*e.cantidad/1000000

                            cantidad = 0

                            for h in computo:
                                if h.proyecto == proyecto and h.tipologia == d.vinculacion:
                                    cantidad = cantidad + h.valor_vacio   

                            valor_capitulo = valor_capitulo + valor_analisis*cantidad

                        else:
                            valor_analisis = 0

                            for e in compo:

                                if e.analisis == d.analisis:

                                    valor_analisis = valor_analisis + e.articulo.valor*e.cantidad/1000000

                            cantidad = 0

                            for h in computo:
                                if h.proyecto == proyecto and h.tipologia == d.vinculacion:
                                    cantidad = cantidad + h.valor_lleno  

                            valor_capitulo = valor_capitulo + valor_analisis*cantidad
        
                    else:

                        valor_analisis = 0

                        for e in compo:

                            if e.analisis == d.analisis:

                                valor_analisis = valor_analisis + e.articulo.valor*e.cantidad/1000000

                        valor_capitulo = valor_capitulo + valor_analisis*float(d.cantidad)

            valor_proyecto = valor_proyecto + valor_capitulo
            
            crudo.append((c, valor_capitulo, 0.0))

        datos =[]

        for i in crudo:
            i = list(i)
            i[2] = i[1]/valor_proyecto*100
            i = tuple(i)
            datos.append(i)

        valor_proyecto_completo = valor_proyecto*1000000

        datos = {"datos":datos, "proyecto":proyecto, "valor_proyecto":valor_proyecto,"valor_proyecto_completo":valor_proyecto_completo}

    
    return render(request, 'presupuestos/presupuestocapitulo.html', {"datos":datos})

# ----------------------------------------------------- VISTAS PARA VER ANALISIS----------------------------------------------

def ver_analisis(request, id_analisis):

    analisis = Analisis.objects.get(codigo = id_analisis)
    composi = CompoAnalisis.objects.all()
    lista_compo = []

    for i in composi:
        if i.analisis == analisis:
            lista_compo.append(i)

    total = 0

    for c in lista_compo:
        total = total + c.articulo.valor*c.cantidad

    lista_final = []

    for d in lista_compo:
        total_renglon = d.cantidad*d.articulo.valor
        inc = (total_renglon/total)*100
        lista_final.append((d, total_renglon, inc))

    datos = {"analisis":analisis, "lista_final":lista_final, "total":total}

    
    return render(request, 'analisis/veranalisis.html', {"datos":datos})


# ----------------------------------------------------- VISTAS PARA LISTAR ANALISIS----------------------------------------------

def analisis_list(request):

    analisis = Analisis.objects.all()
    composicion = CompoAnalisis.objects.all()
    datos = []

    for i in analisis:

        valor = 0

        for c in composicion:

            if i == c.analisis:
                valor = valor +c.articulo.valor*c.cantidad

        datos.append((i, valor))

        #Aqui empieza el filtro

    if request.method == 'POST':

        palabra_buscar = request.POST.items()

        datos_viejos = datos

        datos = []   

        for i in palabra_buscar:

            if i[0] == "palabra":
        
                palabra_buscar = i[1]

        if str(palabra_buscar) == "":

            datos = datos_viejos

        else:
        
            for i in datos_viejos:

                palabra =(str(palabra_buscar))

                buscador = (str(i[0].nombre)+str(i[0].codigo))

                if palabra.lower() in buscador.lower():

                    datos.append(i)


    #Aqui termina el filtro

    return render(request, 'analisis/listaanalisis.html', {"datos":datos})

# ----------------------------------------------------- VISTAS PARA PANEL DE ANALISIS----------------------------------------------

def panelanalisis(request):

    analisis = Analisis.objects.all()
    composicion = CompoAnalisis.objects.all()
    datos = []

    for i in analisis:

        valor = 0

        for c in composicion:

            if i == c.analisis:
                valor = valor +c.articulo.valor*c.cantidad

        datos.append((i, valor))

            #Aqui empieza el filtro

    if request.method == 'POST':

        palabra_buscar = request.POST.items()

        datos_viejos = datos

        datos = []   

        for i in palabra_buscar:

            if i[0] == "palabra":
        
                palabra_buscar = i[1]

        if str(palabra_buscar) == "":

            datos = datos_viejos

        else:
        
            for i in datos_viejos:

                palabra =(str(palabra_buscar))

                buscador = (str(i[0].nombre)+str(i[0].codigo))

                if palabra.lower() in buscador.lower():

                    datos.append(i)


    #Aqui termina el filtro

    return render(request, 'analisis/panelanalisis.html', {"datos":datos})
    
# ----------------------------------------------------- VISTAS PARA CREAR ANALISIS ----------------------------------------------

def crearanalisis(request):

    articulos = Articulos.objects.all()

    mensaje = ""

    datos = {'articulos':articulos, "mensaje":mensaje}

    if request.method == 'POST':

        datos_p = request.POST.items()

        datos_post = []

        for t in datos_p:

            datos_post.append(t)

        tupla = datos_post[1]

        codigo_tupla = tupla[1]

        if len(Analisis.objects.filter(codigo = codigo_tupla)) > 0:

            mensaje = "Este codigo ya se encuentra en la base"

            datos = {'articulos':articulos, "mensaje":mensaje}

        else:

            resto = []

            datos_p = request.POST.items()

            for i in datos_p:

                if i[0] == "codigo":
                    
                    codigo_analisis = i[1]

                elif i[0] == "nombre":
                    
                    nombre = i[1]

                elif i[0] == "unidad":
                    
                    unidad = i[1]
                
                else:

                    resto.append(i)        
                
            id_num = 1

            while Analisis.objects.filter(id = id_num):
                id_num = id_num + 1

           
            b = Analisis(

                id = id_num,
                codigo = codigo_analisis,
                nombre = nombre,
                unidad = unidad,
                )

            b.save()

            valor = 1

            for t in resto:

                if t[0] != "csrfmiddlewaretoken" and valor == 1:

                    valor = 2
    
                    nombre_articulo = t[1]

                elif t[0] != "csrfmiddlewaretoken" and valor == 2:

                    valor = 1

                    cantidad = t[1]

                    datos_compo = CompoAnalisis.objects.all()
                    
                    id_compo = []

                    for c in datos_compo:

                        id_compo.append(c.id)
                        
                    id_num_compo = 1

                    while id_num_compo in id_compo:
                        id_num_compo = id_num_compo + 1
            
                    b = CompoAnalisis(
                        id = id_num_compo,
                        articulo = Articulos.objects.get(nombre=nombre_articulo),
                        analisis = Analisis.objects.get(codigo=codigo_analisis),
                        cantidad = cantidad,
                    )

                    b.save()


            return redirect('Lista de analisis')

    return render(request, 'analisis/crearanalisis.html', {'datos':datos})

# ----------------------------------------------------- VISTAS PARA MODIFICAR ANALISIS ----------------------------------------------

def modificaranalisis(request, id_analisis):

    analisis = Analisis.objects.get(codigo = id_analisis)
    articulos = Articulos.objects.all()
    compo = CompoAnalisis.objects.all()

    datos = []

    for i in compo:

        if i.analisis == analisis:

            datos.append((i.articulo, i.cantidad, i.id))
    

    datos = {"analisis":analisis,
    "articulos":articulos,
    "datos":datos}

    return render(request, 'analisis/modificaranalisis.html', {"datos":datos})



# ----------------------------------------------------- VISTAS PARA PARAMETROS----------------------------------------------

def parametros(request):

    proyectos = Proyectos.objects.all()

    datos = []

    for proyecto in proyectos:

        try:

            parametros = Prametros.objects.get(proyecto = proyecto)
            porc_terreno = parametros.terreno/proyecto.m2*100
            porc_link = parametros.link/proyecto.m2*100
            datos.append((parametros, porc_terreno, porc_link))

        except: 
            print("No esta cargado el parametro de ese proyecto")

    return render(request, 'desde/parametros.html', {'datos': datos})

# ----------------------------------------------------- VISTAS PARA INDICADOR DE PRECIOS----------------------------------------------

def desde(request):

    datos = Desde.objects.all()

    constantes = Constantes.objects.all()

    usd_blue = Constantes.objects.get(nombre = "USD_BLUE")


    for i in datos:

        costo = i.presupuesto.valor

        datos_presupuesto = PresupuestoPorCapitulo(i.presupuesto.proyecto.id)
        datos_saldo = Saldoporcapitulo(i.presupuesto.proyecto.id)

        valor_reposicion = 0

        #Aqui evaluo si se puede usar el presupuesto como valor de costo para calcular el precio min y sugerido

        for p in datos_presupuesto:

            for articulo_cantidad in p[2]:

                valor_reposicion = (valor_reposicion + articulo_cantidad[0].valor*articulo_cantidad[1])

        if valor_reposicion != 0:
            costo = valor_reposicion

        #Aqui calculo el precio min y sugerido

        costo = (costo/(1 + i.parametros.tasa_des_p))*(1 + i.parametros.soft)
        
        costo = costo*(1 + i.parametros.imprevitso)

        aumento_tem = i.parametros.tem_iibb*i.parametros.por_temiibb*(1+i.parametros.ganancia)

        aumento_comer = i.parametros.comer*(1+i.parametros.comer)
        

        costo = costo/(1-aumento_tem- aumento_comer)
        
        m2 = (i.parametros.proyecto.m2 - i.parametros.terreno - i.parametros.link)

        valor_costo = costo/m2
        
        valor_final = valor_costo*(1 + i.parametros.ganancia)

        # Valorizo en dolares el precio de costo y sugerido

        valor_costo_usd = 0

        valor_final_usd = 0

        for c in constantes:

            if str(c.nombre) == 'USD_BLUE':

                valor_costo_usd = valor_costo/c.valor

                valor_final_usd = valor_final/c.valor

        i.valor_costo = valor_costo
        i.valor_costo_usd = valor_costo_usd
        i.valor_final = valor_final
        i.valor_final_usd = valor_final_usd

        i.save()

    # Programa para graficos

    proyectos = Proyectos.objects.all()

    graficos = 0

    datos_pricing = 0
    datos_costo = 0
    datos_sugerido = 0
    proyecto = 0
    ventas_realizadas = 0
    porc_m2 = 0
    porc_no_vendido = 0

    if request.method == 'POST':

        #Trae el proyecto elegido
        proyecto_elegido = request.POST.items()

        #Crea los datos del pricing

        for i in proyecto_elegido:

            if i[0] == "proyecto":
                proyecto = Proyectos.objects.get(id = i[1])

        fechas = []
                
        datos_pricing = PricingResumen.objects.filter(proyecto = proyecto)

        for i in datos_pricing:
            fechas.append(i.fecha)

        #Crea los datos de costo y de sugerido

        datos_costo = []
        datos_sugerido = []

        for fecha in fechas:

            try:
                valor_proyecto = RegistroValorProyecto.objects.get(proyecto = proyecto, fecha = fecha)
                parametros_proyecto = Prametros.objects.get(proyecto = proyecto)

                costo = valor_proyecto.precio_proyecto

                costo = (costo/(1 + parametros_proyecto.tasa_des_p))*(1 + parametros_proyecto.soft)
        
                costo = costo*(1 + parametros_proyecto.imprevitso)

                aumento_tem = parametros_proyecto.tem_iibb*parametros_proyecto.por_temiibb*(1+parametros_proyecto.ganancia)

                aumento_comer = parametros_proyecto.comer*(1+parametros_proyecto.comer)

                costo = costo/(1-aumento_tem- aumento_comer)
                
                m2 = (parametros_proyecto.proyecto.m2 - parametros_proyecto.terreno - parametros_proyecto.link)

                valor_costo = costo/m2 

                datos_costo.append(valor_costo)

                valor_final = valor_costo*(1 + parametros_proyecto.ganancia)

                datos_sugerido.append(valor_final)

            except:
                datos_costo.append("")
                datos_sugerido.append("")

        #Promedio de venta y cantidad

        ventas_realizadas = []

        for fecha in fechas:
            compras_mes = []
            ventas = VentasRealizadas.objects.filter(proyecto = proyecto)
            for venta in ventas:
                if fecha.month == venta.fecha.month and fecha.year == venta.fecha.year and venta.asignacion != "LINK":
                    
                    if venta.anticipo != venta.precio_venta:
                        valor_p_ant = -np.pv(fv=0, rate=(0.82/100), nper=venta.cuotas_pend, pmt=((venta.precio_venta - venta.anticipo)/venta.cuotas_pend))                 
                        valor_m2 = valor_p_ant + venta.anticipo
                        compras_mes.append((valor_m2, venta.m2))
                    else:
                        valor_m2 = venta.precio_venta
                        compras_mes.append((valor_m2, venta.m2))

            cantidad = len(compras_mes)

            total_venta = 0
            total_m2 = 0

            for op in compras_mes:

                total_venta = total_venta + op[0]
                total_m2 = total_m2 + op[1]

            if total_m2 != 0:
                precio_prom = total_venta/total_m2
            else:
                precio_prom = ""

            ventas_realizadas.append((cantidad, precio_prom))
                   
        #Habilito los graficos
        graficos = 1

        # Calcula los m2 vendidos y el perfomance
      
        try:
            ventas = VentasRealizadas.objects.filter(proyecto = proyecto)

            m2_vendidos = 0

            for venta in ventas:

                if venta.asignacion != "LINK":

                    m2_vendidos = m2_vendidos + venta.m2


            porc_m2 = m2_vendidos/proyecto.m2*100
            porc_no_vendido = 100 - porc_m2

        except:

            porc_m2 = 0
            porc_no_vendido = 100
        
    datos = {'datos':datos, 'usd_blue':usd_blue, 
    "proyectos":proyectos, "proyecto":proyecto, 
    "graficos":graficos, "pricing":datos_pricing, 
    "costo":datos_costo, "sugerido":datos_sugerido,
    "ventas":ventas_realizadas,
    "porc_m2":porc_m2,
    "porc_no_vendido":porc_no_vendido}

    return render(request, 'desde/desde.html', {'datos':datos})

# ----------------------------------------------------- VISTAS PARA DATOS DE PROYECTOS ----------------------------------------------

def proyectos(request):

    datos = DatosProyectos.objects.all()

    return render(request, 'datos/projects.html', {'datos':datos})



# --------------------------------> VISTA PARA INFORME PRESUPUESTO <------------------------------------------------------

def InformeArea(request):

    return render(request, 'presupuestos/informearea.html')

# --------------------------------> FUNCIONES Y CLASES USADAS EN LAS VISTAS <------------------------------------------------------

def PresupuestoPorCapitulo(id_proyecto):

    #Modelos que seran necesarios recorrer completos

    proyecto = Proyectos.objects.get(id = id_proyecto)
    capitulo = Capitulos.objects.all()    

    #La lista datos tiene que tener 37 Arrays por cada capitulo

    datos = []
    
    # Vamos a recorrer todos los capitulos y armar una array

    numero_capitulo = 1
    
    for cap in capitulo:

        capitulo = [] 

        modelo = Modelopresupuesto.objects.filter(proyecto = proyecto, capitulo = cap)

        for mod in modelo:

                if mod.cantidad == None:

                    if "SOLO MANO DE OBRA" in str(mod.analisis):

                        computo = Computos.objects.filter(proyecto = proyecto, tipologia = mod.vinculacion)

                        cantidad_computo = 0

                        for comp in computo:

                            cantidad_computo = cantidad_computo + comp.valor_vacio

                        articulos_analisis = CompoAnalisis.objects.filter(analisis = mod.analisis)

                        for compo in articulos_analisis:

                            articulo_cantidad = (compo.articulo, compo.cantidad*cantidad_computo)

                            capitulo.append(articulo_cantidad)


                    else:

                        computo = Computos.objects.filter(proyecto = proyecto, tipologia = mod.vinculacion)

                        cantidad_computo = 0

                        for comp in computo:

                            cantidad_computo = cantidad_computo + comp.valor_lleno

                        articulos_analisis = CompoAnalisis.objects.filter(analisis = mod.analisis)

                        for compo in articulos_analisis:

                            articulo_cantidad = (compo.articulo, compo.cantidad*cantidad_computo)

                            capitulo.append(articulo_cantidad)

    
                else:

                    articulos_analisis = CompoAnalisis.objects.filter(analisis = mod.analisis)

                    for compo in articulos_analisis:

                        articulo_cantidad = (compo.articulo, compo.cantidad*mod.cantidad )

                        capitulo.append(articulo_cantidad)

        datos.append((numero_capitulo, cap, capitulo))

        numero_capitulo += 1


    #Devuelve el numero del capitulo, el nombre y una lista de todos los insumos y la cantidad de cada uno             

    return datos

def Saldoporcapitulo(id_proyecto):

    #Traemos las compras y el presupuesto
    proyecto = Proyectos.objects.get(id = id_proyecto)
    compras = Compras.objects.filter(proyecto = proyecto)
    presupuesto_capitulo = PresupuestoPorCapitulo(id_proyecto)

    #Ordenamos cada capitulo con una lista donde no se repitan los articulos

    datos_viejos = presupuesto_capitulo
    presupuesto_capitulo = []

    contador = 0

    for i in range(37):

        dato = datos_viejos[contador]

        nuevo_art_cant = []

        lista_art_cap = []

        for art_cant in dato[2]:

            lista_art_cap.append(art_cant[0])

        lista_art_cap = list(set(lista_art_cap))
        
        for articulo in lista_art_cap:

            cantidad = 0

            for articulo2 in dato[2]:

                if articulo == articulo2[0]:
                    cantidad = cantidad + articulo2[1]

            if cantidad<0:
                
                print("Hay un articulo negativo completamente")

            nuevo_art_cant.append((articulo, cantidad))

        presupuesto_capitulo.append((dato[0], dato[1], nuevo_art_cant))    
        

        contador += 1

    #Ordenamos la compra para que sea una sola lista

    articulos_comprados = []

    for compra in compras:
        articulos_comprados.append(compra.articulo)

    articulos_comprados = list(set(articulos_comprados))

    #Armamos el stock con todas las compras realizadas de este proyecto

    stock_articulos = []

    for articulo in articulos_comprados:

        compras_articulo = Compras.objects.filter(proyecto = proyecto, articulo = articulo)

        cantidad = 0

        for compra in compras_articulo:
            cantidad = cantidad + compra.cantidad

        stock_articulos.append((articulo, cantidad))

    #Armamos el saldo --> Hay un error ya que al descartar menores a 0, olvidamos que restan consumo

    saldo_capitulo = []

    for capitulo_presupuesto in presupuesto_capitulo:

        articulos_saldo = []

        for articulos_presupuesto in capitulo_presupuesto[2]:

            if articulos_presupuesto[0] in articulos_comprados and articulos_presupuesto[1]>=0:

                contador = 0

                for articulos_stock in stock_articulos:

                    #Si encontramos el articulo del capitulo en el stock, activamos una de las 3 posibilidades

                    if articulos_stock[0] == articulos_presupuesto[0]:

                        articulos_stock = list(articulos_stock)

                        if articulos_stock[1] > articulos_presupuesto[1]:

                            articulos_stock[1] = float(articulos_stock[1]) - float(articulos_presupuesto[1])                           

                            stock_articulos[contador] = list(stock_articulos[contador])
                            stock_articulos[contador][1] = articulos_stock[1]

                            articulos_stock = tuple(articulos_stock)
                            stock_articulos[contador] = tuple(stock_articulos[contador])

                        elif articulos_stock[1] == articulos_presupuesto[1]:

                            articulos_stock[1] = 0
                            stock_articulos[contador] = list(stock_articulos[contador])
                            stock_articulos[contador][1] = articulos_stock[1]

                            articulos_stock = tuple(articulos_stock)
                            stock_articulos[contador] = tuple(stock_articulos[contador])

                        elif articulos_stock[1] < articulos_presupuesto[1]:

                            cantidad_saldo = float(articulos_presupuesto[1]) - float(articulos_stock[1])

                            articulos_stock[1] = 0

                            stock_articulos[contador] = list(stock_articulos[contador])
                            stock_articulos[contador][1] = articulos_stock[1]

                            articulos_saldo.append((articulos_presupuesto[0], cantidad_saldo))

                            articulos_stock = tuple(articulos_stock)
                            stock_articulos[contador] = tuple(stock_articulos[contador])
                    contador += 1
            else:
                articulos_saldo.append(articulos_presupuesto)

        #Modificado con el saldo
                
        saldo_capitulo.append((capitulo_presupuesto[0], capitulo_presupuesto[1], articulos_saldo))


    return saldo_capitulo

class ReporteExplosion(TemplateView):

    def get(self, request, id_proyecto, *args, **kwargs):
        wb = Workbook()

        proyecto = Proyectos.objects.get(id = id_proyecto)
        modelo = Modelopresupuesto.objects.filter(proyecto = proyecto)

        #Con el siguiente conjunto de formulas creamos la explosión de insumos
        
        crudo_analisis = []

        for i in modelo:

            if i.cantidad != None:

                crudo_analisis.append((i.analisis, i.cantidad))

            else:

                if "SOLO MANO DE OBRA" in str(i.analisis.nombre):

                    computos = Computos.objects.filter(tipologia = i.vinculacion, proyecto = proyecto)

                    cantidad = 0

                    for r in computos:
                        cantidad = cantidad + r.valor_vacio

                    crudo_analisis.append((i.analisis, cantidad))

                else:

                    computos = Computos.objects.filter(tipologia = i.vinculacion, proyecto = proyecto)

                    cantidad = 0

                    for r in computos:
                        cantidad = cantidad + r.valor_lleno

                    crudo_analisis.append((i.analisis, cantidad))

        crudo_articulos = []


        for c in crudo_analisis:

            analisis = CompoAnalisis.objects.filter(analisis = c[0])

            for d in analisis:

                cantidad = d.cantidad*c[1]

                crudo_articulos.append((d.articulo, cantidad))

        datos = []

        for t in crudo_articulos:
            datos.append(t[0])

        datos = list(set(datos))

        datos_viejos = datos
        datos = []

        for i in datos_viejos:
            cantidad = 0
            for c in crudo_articulos:
                if i == c[0]:
                    cantidad = cantidad + c[1]
            datos.append((i, cantidad))


        compras = Compras.objects.filter(proyecto = proyecto)

        comprado_aux = ""

        for dato in datos:
            comprado_aux = comprado_aux + str(dato[0])

        datos_viejos = datos
        
        datos = []

        for i in datos_viejos:
            comprado = 0
            for c in compras:
                if c.proyecto == proyecto and c.articulo == i[0]:
                    comprado = comprado + c.cantidad
            
            cantidad_saldo = i[1] - comprado

            saldo = cantidad_saldo * i[0].valor
            
            datos.append((i[0], i[1], comprado, cantidad_saldo, saldo ))

        #Esta parte arma los articulos que no estan en el presupuesto

        mat_no_presup = []

        for compra in compras:
            if str(compra.articulo.nombre) not in comprado_aux and compra.proyecto == proyecto:
                mat_no_presup.append((compra.articulo.nombre, compra.articulo.valor, compra.cantidad))

        cont = 1
        for d in datos:

            if cont == 1:
                ws = wb.active
                ws.title = "Explosion"
                ws["A"+str(cont)] = "CODIGO"
                ws["B"+str(cont)] = "ARTICULO"
                ws["C"+str(cont)] = "UNIDAD"
                ws["D"+str(cont)] = "VALOR"
                ws["E"+str(cont)] = "CANT. PRESPUESTO"
                ws["F"+str(cont)] = "PRESUPUESTO"
                ws["G"+str(cont)] = "COMPRADO"
                ws["H"+str(cont)] = "PENDIENTE"
                ws["I"+str(cont)] = "SALDO PENDIENTE"

                ws["A"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["B"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["C"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["D"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["E"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["F"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["G"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["H"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["I"+str(cont)].alignment = Alignment(horizontal = "center")

                ws["A"+str(cont)].font = Font(bold = True)
                ws["B"+str(cont)].font = Font(bold = True)
                ws["C"+str(cont)].font = Font(bold = True)
                ws["D"+str(cont)].font = Font(bold = True)
                ws["E"+str(cont)].font = Font(bold = True)
                ws["F"+str(cont)].font = Font(bold = True)
                ws["G"+str(cont)].font = Font(bold = True)
                ws["H"+str(cont)].font = Font(bold = True)
                ws["I"+str(cont)].font = Font(bold = True)

                ws.column_dimensions['A'].width = 11.29
                ws.column_dimensions['B'].width = 58.57
                ws.column_dimensions['C'].width = 8.57
                ws.column_dimensions['D'].width = 12.14
                ws.column_dimensions['E'].width = 18.57
                ws.column_dimensions['F'].width = 17.57
                ws.column_dimensions['G'].width = 12
                ws.column_dimensions['H'].width = 11.86
                ws.column_dimensions['I'].width = 17.57

                ws["A"+str(cont+1)] = d[0].codigo
                ws["B"+str(cont+1)] = d[0].nombre
                ws["C"+str(cont+1)] = d[0].unidad
                ws["D"+str(cont+1)] = d[0].valor
                ws["E"+str(cont+1)] = d[1]
                ws["F"+str(cont+1)] = "=D"+str(cont)+"*E"+str(cont)
                ws["G"+str(cont+1)] = d[2]
                ws["H"+str(cont+1)] = d[3]
                ws["I"+str(cont+1)] = d[4]

                ws["A"+str(cont+1)].font = Font(bold = True)
                ws["A"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["D"+str(cont+1)].number_format = '"$"#,##0.00_-'
                ws["C"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["E"+str(cont+1)].number_format = '#,##0.00_-'
                ws["F"+str(cont+1)].number_format = '"$"#,##0.00_-'
                ws["G"+str(cont+1)].number_format = '#,##0.00_-'
                ws["H"+str(cont+1)].number_format = '#,##0.00_-'
                ws["I"+str(cont+1)].font = Font(bold = True)
                ws["I"+str(cont+1)].number_format = '"$"#,##0.00_-'

                cont += 1

            else: 
                ws = wb.active
                ws["A"+str(cont+1)] = d[0].codigo
                ws["B"+str(cont+1)] = d[0].nombre
                ws["C"+str(cont+1)] = d[0].unidad
                ws["D"+str(cont+1)] = d[0].valor
                ws["E"+str(cont+1)] = d[1]
                ws["F"+str(cont+1)] = "=D"+str(cont)+"*E"+str(cont)
                ws["G"+str(cont+1)] = d[2]
                ws["H"+str(cont+1)] = d[3]
                ws["I"+str(cont+1)] = d[4]

                ws["A"+str(cont+1)].font = Font(bold = True)
                ws["A"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["D"+str(cont+1)].number_format = '"$"#,##0.00_-'
                ws["C"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["E"+str(cont+1)].number_format = '#,##0.00_-'
                ws["F"+str(cont+1)].number_format = '"$"#,##0.00_-'
                ws["G"+str(cont+1)].number_format = '#,##0.00_-'
                ws["H"+str(cont+1)].number_format = '#,##0.00_-'
                ws["I"+str(cont+1)].font = Font(bold = True)
                ws["I"+str(cont+1)].number_format = '"$"#,##0.00_-'

                cont += 1
        cont = 1
        for m in mat_no_presup:

            if cont == 1:
                ws = wb.create_sheet('Art-no-pre')
                ws["A"+str(cont)] = "ARTICULO"
                ws["B"+str(cont)] = "VALOR"
                ws["C"+str(cont)] = "CANTIDAD"

                ws["A"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["B"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["C"+str(cont)].alignment = Alignment(horizontal = "center")

                ws["A"+str(cont)].font = Font(bold = True)
                ws["B"+str(cont)].font = Font(bold = True)
                ws["C"+str(cont)].font = Font(bold = True)

                ws.column_dimensions['A'].width = 58.57
                ws.column_dimensions['B'].width = 13
                ws.column_dimensions['C'].width = 13

                cont += 1

            else:
                
                ws = wb['Art-no-pre']
                ws["A"+str(cont)] = m[0]
                ws["B"+str(cont)] = m[1]
                ws["C"+str(cont)] = m[2]

                cont += 1

        #Establecer el nombre del archivo
        nombre_archivo = "Explosion-{0}.xls".format(str(proyecto.nombre))
        #Definir tipo de respuesta que se va a dar
        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response

class ReporteExplosionCap(TemplateView):

    def get(self, request, id_proyecto, *args, **kwargs):
        wb = Workbook()

        #Aqui coloco la formula para calcular

        contador_cap = 0

        for i in range(37):

            saldo = Saldoporcapitulo(id_proyecto)

            datos_viejos = saldo

            datos_saldo = []
            capitulo = []

            contador_cap += 1

            for componentes in datos_viejos:

                if int(componentes[1].id) == int(contador_cap):
                    
                    datos_saldo.append(componentes[2])
                    capitulo.append(componentes[1])

            articulos = []

            for articulo in datos_saldo[0]:
                articulos.append(articulo[0])

            articulos = list(set(articulos))

            articulos_cant = []

            for articulo in articulos:

                cantidad = 0

                for art_can in datos_saldo[0]:

                    if articulo == art_can[0] and art_can[1]>0:
                        cantidad = cantidad + art_can[1]
                articulos_cant.append((articulo, cantidad))
            
            saldo_cap = 0

            datos_viejos = articulos_cant
            datos_saldo = []

            for dato in datos_viejos:
                saldo_cap = saldo_cap + dato[0].valor*dato[1]
                datos_saldo.append((dato[0], dato[1], float(dato[0].valor*dato[1])))

            datos_viejos = datos_saldo
            datos_saldo = []

            for dato in datos_viejos:
                if saldo_cap != 0:
                    inc = float(dato[2])/float(saldo_cap)*100
                else:
                    inc = 0
                datos_saldo.append((dato[0], dato[1], dato[2], inc))


            if len(datos_saldo) == 0:
                datos_saldo = 0

            else:

                datos_saldo.sort(key=lambda tup: tup[3], reverse=True)

                proyecto = Proyectos.objects.get(id = id_proyecto)

                cont = 1
                for d in datos_saldo:

                    if cont == 1:
                        if contador_cap == 1:
                            ws = wb.active
                        else:
                            ws = wb.create_sheet("My sheet")
                        ws.title = "CAP{0}".format(str(contador_cap))
                        ws["A"+str(cont)] = "CODIGO"
                        ws["B"+str(cont)] = "ARTICULO"
                        ws["C"+str(cont)] = "UNIDAD"
                        ws["D"+str(cont)] = "VALOR"
                        ws["E"+str(cont)] = "PENDIENTE"
                        ws["F"+str(cont)] = "SALDO PENDIENTE"
                        ws["G"+str(cont)] = "INC"


                        ws["A"+str(cont)].alignment = Alignment(horizontal = "center")
                        ws["B"+str(cont)].alignment = Alignment(horizontal = "center")
                        ws["C"+str(cont)].alignment = Alignment(horizontal = "center")
                        ws["D"+str(cont)].alignment = Alignment(horizontal = "center")
                        ws["E"+str(cont)].alignment = Alignment(horizontal = "center")
                        ws["F"+str(cont)].alignment = Alignment(horizontal = "center")
                        ws["G"+str(cont)].alignment = Alignment(horizontal = "center")


                        ws["A"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                        ws["A"+str(cont)].fill =  PatternFill("solid", fgColor= "159ABB")
                        ws["B"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                        ws["B"+str(cont)].fill =  PatternFill("solid", fgColor= "159ABB")
                        ws["C"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                        ws["C"+str(cont)].fill =  PatternFill("solid", fgColor= "159ABB")
                        ws["D"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                        ws["D"+str(cont)].fill =  PatternFill("solid", fgColor= "159ABB")
                        ws["E"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                        ws["E"+str(cont)].fill =  PatternFill("solid", fgColor= "159ABB")
                        ws["F"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                        ws["F"+str(cont)].fill =  PatternFill("solid", fgColor= "159ABB")
                        ws["G"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                        ws["G"+str(cont)].fill =  PatternFill("solid", fgColor= "159ABB")


                        ws.column_dimensions['A'].width = 11.29
                        ws.column_dimensions['B'].width = 58.57
                        ws.column_dimensions['C'].width = 8.57
                        ws.column_dimensions['D'].width = 12.14
                        ws.column_dimensions['E'].width = 18.57
                        ws.column_dimensions['F'].width = 17.57
                        ws.column_dimensions['G'].width = 12

                        ws["A"+str(cont+1)] = d[0].codigo
                        ws["B"+str(cont+1)] = d[0].nombre
                        ws["C"+str(cont+1)] = d[0].unidad
                        ws["D"+str(cont+1)] = d[0].valor
                        ws["E"+str(cont+1)] = d[1]
                        ws["F"+str(cont+1)] = d[2]
                        ws["G"+str(cont+1)] = d[3]


                        ws["A"+str(cont+1)].font = Font(bold = True)
                        ws["B"+str(cont+1)].alignment = Alignment(horizontal = "center")
                        ws["C"+str(cont+1)].number_format = '"$"#,##0.00_-'
                        ws["D"+str(cont+1)].alignment = Alignment(horizontal = "center")
                        ws["E"+str(cont+1)].number_format = '#,##0.00_-'
                        ws["F"+str(cont+1)].number_format = '"$"#,##0.00_-'
                        ws["G"+str(cont+1)].number_format = '#,##0.00_-"%"'

                        cont += 1

                    else:
                        if contador_cap == 1: 
                            ws = wb.active
                        else:
                            ws = wb["CAP{0}".format(str(contador_cap))]
                        ws["A"+str(cont+1)] = d[0].codigo
                        ws["B"+str(cont+1)] = d[0].nombre
                        ws["C"+str(cont+1)] = d[0].unidad
                        ws["D"+str(cont+1)] = d[0].valor
                        ws["E"+str(cont+1)] = d[1]
                        ws["F"+str(cont+1)] = d[2]
                        ws["G"+str(cont+1)] = d[3]


                        ws["A"+str(cont+1)].font = Font(bold = True)
                        ws["B"+str(cont+1)].alignment = Alignment(horizontal = "center")
                        ws["C"+str(cont+1)].number_format = '"$"#,##0.00_-'
                        ws["D"+str(cont+1)].alignment = Alignment(horizontal = "center")
                        ws["E"+str(cont+1)].number_format = '#,##0.00_-'
                        ws["F"+str(cont+1)].number_format = '"$"#,##0.00_-'
                        ws["G"+str(cont+1)].number_format = '#,##0.00_-"%"'
        

                        cont += 1

        #Establecer el nombre del archivo
        nombre_archivo = "ExplosionCap-{0}.xls".format(str(proyecto.nombre))
        #Definir tipo de respuesta que se va a dar
        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


def debugsaldo(id_proyecto):

    #Traemos las compras y el presupuesto
    proyecto = Proyectos.objects.get(id = id_proyecto)
    compras = Compras.objects.filter(proyecto = proyecto)
    presupuesto_capitulo = PresupuestoPorCapitulo(id_proyecto)

    #Ordenamos cada capitulo con una lista donde no se repitan los articulos

    datos_viejos = presupuesto_capitulo
    presupuesto_capitulo = []

    contador = 0

    mensaje = []

    for i in range(37):

        dato = datos_viejos[contador]

        lista_art_cap = []

        for art_cant in dato[2]:

            lista_art_cap.append(art_cant[0])

        lista_art_cap = list(set(lista_art_cap))
        
        for articulo in lista_art_cap:

            cantidad = 0

            for articulo2 in dato[2]:

                if articulo == articulo2[0]:
                    cantidad = cantidad + articulo2[1]

            if cantidad<0:
            

                mensaje.append((articulo.nombre, dato[1]))  
        

        contador += 1

        return mensaje
    












